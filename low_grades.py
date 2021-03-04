from openpyxl import load_workbook
# from pathlib import Path # Use when on local machine

# Excel gradebook name
gb_name = 'Test.xlsx'
# Load the gradebook
wb = load_workbook(gb_name)
# get the active worksheet (there should only be 1 even if multiple class exports)(multiple class exports are not supported)
ws = wb.active

def get_assignments(): # and max scores
	# looks in row 5 until there are 5 None values in a row (for row 5)
	assignment_names = [] # Initialize list
	max_scores = [] # Initialize list
	# We need to store the first 5 rows for this loop to work
	for col in range(7,12):
		assignment_names.append(ws.cell(row = 5, column = col).value)
		max_scores.append(ws.cell(row = 6, column = col).value) # might as well do these while I am here
	n = 5 # initialize assignment count
	while (assignment_names[n-5] != None) or (assignment_names[n-4] != None) or (assignment_names[n-3] != None) or (assignment_names[n-2] != None) or (assignment_names[n-1] != None): # past 5 assignments cannot be 'None'
		assignment_names.append(ws.cell(row = 5, column = 8 + n).value)
		max_scores.append(ws.cell(row = 6, column = 8 + n).value)
		n += 1
	return [assignment_names[0:len(assignment_names)-5], max_scores[0:len(assignment_names)-5]]

assignments = get_assignments()[0]
max_scores = get_assignments()[1]
tot_assignments = len(assignments) # not really the number of assignments because some columns are blank

def get_student_grades(student, tot_assignments):
	# student represents column of the student to get grades for
	# makes a list of all assignment grades for that student
	student_grades = [] # initialize list
	student_grades.append(ws.cell(row = student, column = 1).value) # add the student name before the grades
	for grade in range(tot_assignments):
		student_grades.append(ws.cell(row = student, column = 8 + grade).value)
	return student_grades

def get_n_students():
	# Looks from A8 down until it sees the '#' symbol and then stops
	student = 'student' # initialize list
	row = 8 # Aeries default location for the first student's name
	while (student != '#'): # assumes no more than 50 students
		student = ws.cell(row = row, column = 1).value
		row += 1
	return row - 9

n_students = get_n_students()

# I might have this export to my own gradebook in the future...thus adding headers for rows and columns
# Build Clean Gradebook Headers
headers = [] # Eventually look like 'Student Name' 'Assignment 1' 'Assignment 2' . . .
headers.append('Student Name')
for val in assignments:
	headers.append(val)
# Need a title for the row of max scores
max_scores.insert(0,'Max Score')

# Create blank list/matrix for a clean gradebook
gb_clean = []
# Append headers and max scores
gb_clean.append(headers)
gb_clean.append(max_scores)
# Add all student grades to the Clean Gradebook
for student in range(n_students):
	gb_clean.append(get_student_grades(8 + student, tot_assignments))
# Delete columns with 'None' for assignment names (Up to 4 of them)
# First find the columns with the none and store them in 'bad'
bad = [] # initialize list of bad columns (aka columns that have None for the assignment name)
for x in range(len(gb_clean[0])):
	if gb_clean[0][x] == None:
		bad.append(x)
# Delete each of the bad columns
# Delete from the end instead of the beginning, otherwise index messed up
for i in sorted(bad, reverse = True):
	for row in range(len(gb_clean)):
		del gb_clean[row][i]

# Don't worry about the following "grades":
# None means the grade has not been entered
# 'TX' means the grade is temporarily excused

def isfloat(value):
  try:
    float(value)
    return True
  except ValueError:
    return False

def get_low_grades(assignment_numbers, thresholds):
	# if grade for the assignment number is less than the threshold percent, add to the matrix
	# assignment_numbers needs to be a list (even if it is a list with 1 element)
	# list will have student, assignment, student, assignment, etc
	low_grades = [] # Initialize matrix
	# now get low scores
	# this creates a list for each student with low assignment information
	# list of lists: [student_name, [assignment1_name, assignment1_score], [assignment2_name, assignment2_score]]
	for student in range(2, len(gb_clean)):
		low_student_grades = [] # Initialize list for just 1 student
		low_student_grades.append(gb_clean[student][0]) # empty lists (no low assignments) will have to be deleted after this for loop
		for num in range(len(assignment_numbers)):
			try:
				if float(gb_clean[student][assignment_numbers[num]]) / float(gb_clean[1][assignment_numbers[num]]) < thresholds[num] / 100:
					assignment_data = [] # Initialize
					assignment_data.append(gb_clean[0][assignment_numbers[num]]) # assignment name
					assignment_data.append(float(gb_clean[student][assignment_numbers[num]]) / float(gb_clean[1][assignment_numbers[num]]) * 100) # percent
					low_student_grades.append(assignment_data) # add to list for 1 student
			except:
				print(f'\'{gb_clean[student][0]}\' does not have a numeric score for {gb_clean[0][assignment_numbers[num]]}.\nSkipping Student.')
		low_grades.append(low_student_grades) # add to the list for all students
	return low_grades